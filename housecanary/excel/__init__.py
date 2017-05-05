"""Module for creating Excel exports of HouseCanary API data"""

from __future__ import print_function
import os
import csv
import time
import io
import sys
import openpyxl
from builtins import str
from slugify import slugify
from . import analytics_data_excel
from . import utilities
from .. import ApiClient
from .. import exceptions


def export_analytics_data_to_excel(data, output_file_name, result_info_key,
                                   identifier_keys, extra_identifiers):
    """Creates an Excel file containing data returned by the Analytics API

    Args:
        data: Analytics API data as a list of dicts
        output_file_name: File name for output Excel file (use .xlsx extension).

    """
    workbook = create_excel_workbook(data, result_info_key, identifier_keys, extra_identifiers)
    workbook.save(output_file_name)
    print('Saved Excel file to {}'.format(output_file_name))


def export_analytics_data_to_csv(data, output_folder, result_info_key,
                                 identifier_keys, extra_identifiers):
    """Creates CSV files containing data returned by the Analytics API.
       Creates one file per requested endpoint and saves it into the
       specified output_folder

    Args:
        data: Analytics API data as a list of dicts
        output_folder: Path to a folder to save the CSV files into
    """
    workbook = create_excel_workbook(data, result_info_key, identifier_keys, extra_identifiers)

    suffix = '.csv'

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for worksheet in workbook.worksheets:
        file_name = utilities.convert_title_to_snake_case(worksheet.title)

        file_path = os.path.join(output_folder, file_name + suffix)

        mode = 'w'
        if sys.version_info[0] < 3:
            mode = 'wb'
        with io.open(file_path, mode) as output_file:
            csv_writer = csv.writer(output_file)
            for row in worksheet.rows:
                csv_writer.writerow([cell.value for cell in row])

    print('Saved CSV files to {}'.format(output_folder))


def concat_excel_reports(addresses, output_file_name, endpoint, report_type,
                         retry, api_key, api_secret, files_path):
    """Creates an Excel file made up of combining the Value Report or Rental Report Excel
       output for the provided addresses.

    Args:
        addresses: A list of (address, zipcode) tuples
        output_file_name: A file name for the Excel output
        endpoint: One of 'value_report' or 'rental_report'
        report_type: One of 'full' or 'summary'
        retry: optional boolean to retry if rate limit is reached
        api_key: optional API Key
        api_secret: optional API Secret
        files_path: Path to save individual files. If None, don't save files
    """
    # create the master workbook to output
    master_workbook = openpyxl.Workbook()

    if api_key is not None and api_secret is not None:
        client = ApiClient(api_key, api_secret)
    else:
        client = ApiClient()

    errors = []

    # for each address, call the API and load the xlsx content in a workbook.
    for index, addr in enumerate(addresses):
        print('Processing {}'.format(addr['address']))
        result = _get_excel_report(
            client, endpoint, addr['address'], addr['zipcode'], report_type, retry)

        if not result['success']:
            print('Error retrieving report for {}'.format(addr['address']))
            print(result['content'])
            errors.append({'address': addr['address'], 'message': result['content']})
            continue

        orig_wb = openpyxl.load_workbook(filename=io.BytesIO(result['content']))

        _save_individual_file(orig_wb, files_path, addr['address'])

        # for each worksheet for this address
        for sheet_name in orig_wb.get_sheet_names():
            # if worksheet doesn't exist in master workbook, create it
            if sheet_name in master_workbook.get_sheet_names():
                master_ws = master_workbook.get_sheet_by_name(sheet_name)
            else:
                master_ws = master_workbook.create_sheet(sheet_name)

            # get all the rows in the address worksheet
            orig_rows = orig_wb.get_sheet_by_name(sheet_name).rows

            if sheet_name == 'Summary' or sheet_name == 'Chart Data':
                _process_non_standard_sheet(master_ws, orig_rows, addr, index)
                continue

            _process_standard_sheet(master_ws, orig_rows, addr, index)

    # remove the first sheet which will be empty
    master_workbook.remove(master_workbook.worksheets[0])

    # if any errors occurred, write them to an "Errors" worksheet
    if len(errors) > 0:
        errors_sheet = master_workbook.create_sheet('Errors')
        for error_idx, error in enumerate(errors):
            errors_sheet.cell(row=error_idx+1, column=1, value=error['address'])
            errors_sheet.cell(row=error_idx+1, column=2, value=error['message'])

    # save the master workbook to output_file_name
    adjust_column_width_workbook(master_workbook)
    output_file_path = os.path.join(files_path, output_file_name)
    master_workbook.save(output_file_path)
    print('Saved output to {}'.format(output_file_path))


def _process_standard_sheet(master_ws, orig_rows, addr, address_index):
    # if this is the first address, add headers for identifier columns passed into the input
    if address_index == 0:
        for idx, key in enumerate(addr.keys()):
            master_ws.cell(row=1, column=idx+1, value=key)

    # get the next row in the master worksheet to start writing to.
    # this actually sets the next row to the last row with values in it,
    # but that's good because the first row of the next address sheet
    # is skipped in order to skip the header.
    next_row_idx = 1 if address_index == 0 else master_ws.max_row

    # go through the rows from the address worksheet
    for orig_row_idx, orig_row in enumerate(orig_rows):
        if address_index > 0 and orig_row_idx == 0:
            # after the first address, skip the header rows
            continue

        # write the header columns
        if orig_row_idx > 0:
            for idx, header_val in enumerate(addr.values()):
                master_ws.cell(row=next_row_idx + orig_row_idx, column=idx+1, value=header_val)

        start_cell_idx = len(addr.keys())

        # copy over the address sheet's cells
        # starting at the row we left off at
        _copy_row_to_worksheet(master_ws, orig_row, next_row_idx, orig_row_idx, start_cell_idx)


def _process_non_standard_sheet(master_ws, orig_rows, addr, address_index):
    # for the Summary sheet, there are multiple rows with different data,
    # so we'll simply copy the rows as they are

    if address_index == 0:
        # add headers for identifier columns passed into the input
        for idx, key in enumerate(addr.keys()):
            master_ws.cell(row=1, column=idx+1, value=key)

    # first, let's get the next row to write to,
    # leaving a space between the data from the previous address
    next_row_idx = 2 if address_index == 0 else master_ws.max_row + 2

    # add identifier values passed into the input
    for idx, header_val in enumerate(addr.values()):
        master_ws.cell(row=next_row_idx, column=idx+1, value=header_val)

    start_cell_idx = len(addr.keys())

    for orig_row_idx, orig_row in enumerate(orig_rows):
        # copy over the address sheet's cells
        # starting at the row we left off at
        _copy_row_to_worksheet(master_ws, orig_row, next_row_idx, orig_row_idx, start_cell_idx)


def _copy_row_to_worksheet(master_ws, orig_row, next_row_idx, orig_row_idx, start_cell_idx):
    for orig_cell_idx, orig_cell in enumerate(orig_row):
        master_ws.cell(
            row=next_row_idx + orig_row_idx,
            column=orig_cell_idx + start_cell_idx + 1,  # start_cell_idx accounts for identifier columns
            value=orig_cell.value
        )


def _get_excel_report(client, endpoint, address, zipcode, report_type, retry):
    if retry:
        while True:
            try:
                return _make_report_request(client, endpoint, address, zipcode, report_type)
            except exceptions.RateLimitException as e:
                rate_limit = e.rate_limits[0]
                utilities.print_rate_limit_error(rate_limit)

                if rate_limit['reset_in_seconds'] >= 300:
                    # Rate limit will take more than 5 minutes to reset, so just fail
                    return {'success': False, 'content': str(e)}

                print('Will retry once rate limit resets...')
                time.sleep(rate_limit['reset_in_seconds'])
            except exceptions.RequestException as e:
                return {'success': False, 'content': str(e)}

    # otherwise just try once.
    try:
        return _make_report_request(client, endpoint, address, zipcode, report_type)
    except exceptions.RequestException as e:
        return {'success': False, 'content': str(e)}


def _make_report_request(client, endpoint, address, zipcode, report_type):
    if endpoint == 'rental_report':
        response = client.property.rental_report(address, zipcode, 'xlsx')
    else:
        response = client.property.value_report(address, zipcode, report_type, 'xlsx')
    return {'success': True, 'content': response.content}


def _save_individual_file(workbook, files_path, addr):
    if not os.path.exists(files_path):
        os.makedirs(files_path)

    file_name = slugify('{}-{}'.format(addr, time.strftime('%Y-%m-%d_%H-%M-%S')))
    file_path = os.path.join(files_path, '{}.xlsx'.format(file_name))

    workbook.save(file_path)
    print('Saved output to {}'.format(file_path))


def create_excel_workbook(data, result_info_key, identifier_keys, extra_identifiers):
    """Calls the analytics_data_excel module to create the Workbook"""
    workbook = analytics_data_excel.get_excel_workbook(
        data, result_info_key, identifier_keys, extra_identifiers)
    adjust_column_width_workbook(workbook)
    return workbook


def adjust_column_width_workbook(workbook):
    """Adjust column width for all sheets in workbook."""
    for worksheet in workbook.worksheets:
        adjust_column_width(worksheet)


def adjust_column_width(worksheet):
    """Adjust column width in worksheet.

    Args:
        worksheet: worksheet to be adjusted
    """
    dims = {}
    padding = 1
    for row in worksheet.rows:
        for cell in row:
            if not cell.value:
                continue
            dims[cell.column] = max(
                dims.get(cell.column, 0),
                len(str(cell.value))
            )
    for col, value in list(dims.items()):
        worksheet.column_dimensions[col].width = value + padding
