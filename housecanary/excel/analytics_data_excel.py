import copy
import openpyxl

from . import utilities


KEY_TO_WORKSHEET_MAP = {
    'Msa Details': 'MSA Details'
}

LEADING_COLUMNS = (
    'address',
    'unit',
    'city',
    'state',
    'zipcode',
    'slug',
    'block_id',
    'msa',
    'num_bins',
    'property_type',
    'client_value',
    'client_value_sqft',
    'meta'
)

LEADING_WORKSHEETS = ()


def get_excel_workbook(api_data, result_info_key, identifier_keys):
    """Generates an Excel workbook object given api_data returned by the Analytics API

    Args:
        api_data: Analytics API data as a list of dicts (one per identifier)
        result_info_key: the key in api_data dicts that contains the data results
        identifier_keys: the list of keys used as requested identifiers
                         (address, zipcode, block_id, etc)

    Returns:
        raw excel file data
    """

    cleaned_data = []

    for item_data in api_data:
        result_info = item_data.pop(result_info_key, {})

        cleaned_item_data = {}

        if 'meta' in item_data:
            meta = item_data.pop('meta')
            cleaned_item_data['meta'] = meta

        for key in item_data:
            cleaned_item_data[key] = item_data[key]['result']

        cleaned_item_data[result_info_key] = result_info

        cleaned_data.append(cleaned_item_data)

    data_list = copy.deepcopy(cleaned_data)

    workbook = openpyxl.Workbook()

    write_worksheets(workbook, data_list, result_info_key, identifier_keys)

    return workbook


def write_worksheets(workbook, data_list, result_info_key, identifier_keys):
    """Writes rest of the worksheets to workbook.

    Args:
        workbook: workbook to write into
        data_list: Analytics API data as a list of dicts
        result_info_key: the key in api_data dicts that contains the data results
        identifier_keys: the list of keys used as requested identifiers
                         (address, zipcode, block_id, etc)
    """

    # we can use the first item to figure out the worksheet keys
    worksheet_keys = get_worksheet_keys(data_list[0], result_info_key)

    for key in worksheet_keys:

        title = key.split('/')[1]

        title = utilities.convert_snake_to_title_case(title)

        title = KEY_TO_WORKSHEET_MAP.get(title, title)

        if key == 'property/nod':
            # the property/nod endpoint needs to be split into two worksheets
            create_property_nod_worksheets(workbook, data_list, result_info_key, identifier_keys)
        else:
            # all other endpoints are written to a single worksheet

            # Maximum 31 characters allowed in sheet title
            worksheet = workbook.create_sheet(title=title[:31])

            processed_data = process_data(key, data_list, result_info_key, identifier_keys)

            write_data(worksheet, processed_data)

    # remove the first, unused empty sheet
    workbook.remove_sheet(workbook.active)


def create_property_nod_worksheets(workbook, data_list, result_info_key, identifier_keys):
    """Creates two worksheets out of the property/nod data because the data
       doesn't come flat enough to make sense on one sheet.

       Args:
            workbook: the main workbook to add the sheets to
            data_list: the main list of data
            result_info_key: the key in api_data dicts that contains the data results
                             Should always be 'address_info' for property/nod
            identifier_keys: the list of keys used as requested identifiers
                            (address, zipcode, city, state, etc)
    """
    nod_details_list = []
    nod_default_history_list = []

    for prop_data in data_list:
        nod_data = prop_data['property/nod']

        if nod_data is None:
            nod_data = {}

        default_history_data = nod_data.pop('default_history', [])

        _set_identifier_fields(nod_data, prop_data, result_info_key, identifier_keys)

        nod_details_list.append(nod_data)

        for item in default_history_data:
            _set_identifier_fields(item, prop_data, result_info_key, identifier_keys)
            nod_default_history_list.append(item)

    worksheet = workbook.create_sheet(title='NOD Details')
    write_data(worksheet, nod_details_list)

    worksheet = workbook.create_sheet(title='NOD Default History')
    write_data(worksheet, nod_default_history_list)


def get_worksheet_keys(data_dict, result_info_key):
    """Gets sorted keys from the dict, ignoring result_info_key and 'meta' key
    Args:
        data_dict: dict to pull keys from

    Returns:
        list of keys in the dict other than the result_info_key
    """
    keys = set(data_dict.keys())
    keys.remove(result_info_key)
    if 'meta' in keys:
        keys.remove('meta')
    return sorted(keys)


def get_keys(data_list, leading_columns=LEADING_COLUMNS):
    """Gets all possible keys from a list of dicts, sorting by leading_columns first

    Args:
        data_list: list of dicts to pull keys from
        leading_columns: list of keys to put first in the result

    Returns:
        list of keys to be included as columns in excel worksheet
    """
    all_keys = set().union(*(list(d.keys()) for d in data_list))

    leading_keys = []

    for key in leading_columns:
        if key not in all_keys:
            continue
        leading_keys.append(key)
        all_keys.remove(key)

    return leading_keys + sorted(all_keys)


def write_data(worksheet, data):
    """Writes data into worksheet.

    Args:
        worksheet: worksheet to write into
        data: data to be written
    """
    if not data:
        return

    if isinstance(data, list):
        rows = data
    else:
        rows = [data]

    if isinstance(rows[0], dict):
        keys = get_keys(rows)
        worksheet.append([utilities.convert_snake_to_title_case(key) for key in keys])
        for row in rows:
            values = [get_value_from_row(row, key) for key in keys]
            worksheet.append(values)
    elif isinstance(rows[0], list):
        for row in rows:
            values = [utilities.normalize_cell_value(value) for value in row]
            worksheet.append(values)
    else:
        for row in rows:
            worksheet.append([utilities.normalize_cell_value(row)])


def get_value_from_row(row, key):
    """Gets a value and normalizes it's value"""
    if key in row:
        return utilities.normalize_cell_value(row[key])
    return ''


def process_data(key, data_list, result_info_key, identifier_keys):
    """ Given a key as the endpoint name, pulls the data for that endpoint out
        of the data_list for each address, processes the data into a more
        excel-friendly format and returns that data.

        Args:
            key: the endpoint name of the data to process
            data_list: the main data list to take the data from
            result_info_key: the key in api_data dicts that contains the data results
            identifier_keys: the list of keys used as requested identifiers
                             (address, zipcode, block_id, etc)

        Returns:
            A list of dicts (rows) to be written to a worksheet
    """
    master_data = []

    for item_data in data_list:
        data = item_data[key]

        if data is None:
            current_item_data = {}
        else:
            if key == 'property/value':
                current_item_data = data['value']

            elif key == 'property/details':
                top_level_keys = ['property', 'assessment']
                current_item_data = flatten_top_level_keys(data, top_level_keys)

            elif key == 'property/school':
                current_item_data = data['school']

                school_list = []
                for school_type_key in current_item_data:
                    schools = current_item_data[school_type_key]
                    for school in schools:
                        school['school_type'] = school_type_key
                        school['school_address'] = school['address']
                        school['school_zipcode'] = school['zipcode']
                        school_list.append(school)

                current_item_data = school_list

            elif key == 'property/value_forecast':
                current_item_data = {}
                for month_key in data:
                    current_item_data[month_key] = data[month_key]['value']

            elif key in ['property/value_within_block', 'property/rental_value_within_block']:
                current_item_data = flatten_top_level_keys(data, [
                    'housecanary_value_percentile_range',
                    'housecanary_value_sqft_percentile_range',
                    'client_value_percentile_range',
                    'client_value_sqft_percentile_range'
                ])

            elif key in ['property/zip_details', 'zip/details']:
                top_level_keys = ['multi_family', 'single_family']
                current_item_data = flatten_top_level_keys(data, top_level_keys)

            else:
                current_item_data = data

        if isinstance(current_item_data, dict):
            _set_identifier_fields(current_item_data, item_data, result_info_key, identifier_keys)

            master_data.append(current_item_data)
        else:
            # it's a list
            for item in current_item_data:
                _set_identifier_fields(item, item_data, result_info_key, identifier_keys)

            master_data.extend(current_item_data)

    return master_data


def _set_identifier_fields(item, item_data, result_info_key, identifier_keys):
    result_info = item_data[result_info_key]
    for k in identifier_keys:
        if k == 'meta':
            item['meta'] = item_data['meta']

        # skip the special query param keys since they are not always returned in the result info
        elif k not in ['client_value', 'client_value_sqft', 'num_bins', 'property_type']:
            item[k] = result_info[k]


def flatten_top_level_keys(data, top_level_keys):
    """ Helper method to flatten a nested dict of dicts (one level)

        Example:
            {'a': {'b': 'bbb'}} becomes {'a_-_b': 'bbb'}

            The separator '_-_' gets formatted later for the column headers

        Args:
            data: the dict to flatten
            top_level_keys: a list of the top level keys to flatten ('a' in the example above)
    """
    flattened_data = {}

    for top_level_key in top_level_keys:
        if data[top_level_key] is None:
            flattened_data[top_level_key] = None
        else:
            for key in data[top_level_key]:
                flattened_data['{}_-_{}'.format(top_level_key, key)] = data[top_level_key][key]

    return flattened_data
